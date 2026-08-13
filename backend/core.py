import re
import os
import difflib
import pandas as pd
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from datetime import datetime
from database import ColumnMapping, BrandRule, CategorySpecTemplate, UserCorrection, ProcessingHistory

# --- UTILITIES & PARSERS ---

def slugify(text: str) -> str:
    """
    Slugify product titles to create clean, unique Shopify Handles,
    preserving casing and other characters exactly as in historical listings.
    """
    if not text:
        return ""
    # Strip leading/trailing whitespaces
    s = str(text).strip()
    # Replace spaces and slashes with hyphens
    s = re.sub(r'[\s/]+', '-', s)
    # Replace multiple hyphens with a single hyphen
    s = re.sub(r'-+', '-', s)
    return s.strip('-')

def read_excel_fast(path: str, sheet_name: str, max_rows: int = None) -> pd.DataFrame:
    """
    Fast streaming Excel reader using openpyxl read_only mode.
    Correctly handles column de-duplication and strips whitespace.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in Excel file.")
    
    sheet = wb[sheet_name]
    data = []
    rows = sheet.iter_rows(values_only=True)
    
    try:
        headers = next(rows)
    except StopIteration:
        return pd.DataFrame()
        
    headers = [str(h).strip() if h is not None else f"Col{idx}" for idx, h in enumerate(headers)]
    
    # De-duplicate headers
    unique_headers = []
    seen = {}
    for h in headers:
        if h in seen:
            seen[h] += 1
            unique_headers.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            unique_headers.append(h)
            
    for i, r in enumerate(rows):
        if max_rows and i >= max_rows:
            break
        r_list = list(r) if r is not None else []
        # Pad or slice to match headers length
        if len(r_list) < len(unique_headers):
            r_list += [None] * (len(unique_headers) - len(r_list))
        elif len(r_list) > len(unique_headers):
            r_list = r_list[:len(unique_headers)]
            
        r_list = [str(x).strip() if x is not None else "" for x in r_list]
        data.append(r_list)
        
    df = pd.DataFrame(data, columns=unique_headers)
    return df.reset_index(drop=True)

def read_csv_fast(path: str) -> pd.DataFrame:
    """
    Reads a CSV file with robust encoding fallbacks (utf-8, cp1252, etc.).
    """
    for enc in ['utf-8', 'utf-8-sig', 'cp1252', 'latin-1']:
        try:
            df = pd.read_csv(path, encoding=enc, dtype=str, keep_default_na=False)
            return df.reset_index(drop=True)
        except Exception:
            continue
    raise ValueError(f"Could not read CSV file '{os.path.basename(path)}' with any supported encoding.")

# --- SEMANTIC AUTO-MAPPING ENGINE ---

SEMANTIC_THESAURUS = {
    "Title": ["d2c title", "amazon title", "myntra title", "nykaa title", "pdp name", "title", "product name", "item description"],
    "Body (HTML)": ["description", "body", "bullet points", "marketing content", "product description"],
    "Variant SKU": ["item code", "sku", "barcode", "variant sku", "variant barcode", "sno"],
    "Variant Barcode": ["item name", "style", "style code", "item color", "color", "shade name"],
    "Option1 Value": ["size", "option1 value", "euro", "length in cm", "age group"],
    "Variant Compare At Price": ["mrp", "compare price", "list price", "compare at price", "variant compare at price"],
    "Variant Price": ["price", "selling price", "variant price", "d2c price"],
    "Upper Material": ["upper", "material", "fabric", "upper material"],
    "Sole Material": ["sole", "sole material"]
}

def auto_map_columns(source_columns: list, source_sheet_type: str, db: Session) -> dict:
    """
    Auto-maps column headers in uploaded files using user-corrections and fuzzy semantic ratios.
    Returns dictionary: { shopify_target_header: source_header }
    """
    mappings = {}
    
    # 1. Fetch user-approved corrections from SQLite
    corrections = db.query(UserCorrection).filter(UserCorrection.source_sheet == source_sheet_type).all()
    correction_dict = {c.user_column: c.source_column for c in corrections}
    
    # 2. Fetch default mappings
    default_maps = db.query(ColumnMapping).filter(ColumnMapping.source_sheet == source_sheet_type).all()
    default_dict = {m.shopify_column: m.source_column for m in default_maps}
    
    for shopify_col, synonyms in SEMANTIC_THESAURUS.items():
        # A. Use user correction if exists
        if shopify_col in correction_dict and correction_dict[shopify_col] in source_columns:
            mappings[shopify_col] = correction_dict[shopify_col]
            continue
            
        # B. Use default mapped seed column if exists
        if shopify_col in default_dict and default_dict[shopify_col] in source_columns:
            mappings[shopify_col] = default_dict[shopify_col]
            continue
            
        # C. Fuzzy match using ratio scoring
        best_match = None
        best_score = 0.0
        
        for col in source_columns:
            col_lower = col.lower()
            # Exact match
            if col_lower == shopify_col.lower():
                best_match = col
                best_score = 1.0
                break
                
            # Synonym exact match
            if col_lower in synonyms:
                best_match = col
                best_score = 0.95
                break
                
            # Fuzzy match synonyms
            for syn in synonyms:
                score = difflib.SequenceMatcher(None, col_lower, syn).ratio()
                if score > best_score:
                    best_score = score
                    best_match = col
                    
        # Apply fuzzy match if above threshold (e.g. 70% match)
        if best_match and best_score >= 0.70:
            mappings[shopify_col] = best_match
            
    return mappings

# --- CATEGORY RESOLVER & SUBCATEGORY MAPPER ---

DENIM_SUBCATEGORY_MAP = {
    "JEANS": "Jeans",
    "SHORTS": "Shorts",
    "SKIRT": "Skirt",
    "JACKET": "Jacket",
    "SHIRT F/S": "Shirt",
    "SHIRT H/S": "Shirt",
    "SHIRT": "Shirt",
    "TOP": "Top",
    "DENIM": "Denim",
    "DUNGAREE": "Dungaree",
    "DUNGAREE SET F/S": "Dungaree",
    "DUNGAREE SET": "Dungaree",
    "BERMUDA": "Bermuda",
    "JUMPSUIT": "Jumpsuit",
    "DRESS": "Dress",
    "CO-ORD SET": "Clothing Set",
    "COORD SET": "Clothing Set",
    "CO-ORD": "Clothing Set",
    "CLOTHING SET": "Clothing Set"
}

def resolve_category(category: str, sub_category: str = "") -> str:
    """
    Resolves and normalizes category names based on catalog business rules:
    1. 'POLO T-SHIRT' is strictly converted to 'T-Shirt'.
    2. 'DENIM' is converted to its Subcategory, with special mappings for Shirts, Dungarees, Sets, etc.
    3. Normalizes other specific category names (Soft Toys, Sandals, Baby Booties).
    """
    cat_raw = str(category).strip()
    cat_upper = cat_raw.upper()
    subcat_upper = str(sub_category).strip().upper()
    
    # 1. Strictly convert POLO T-SHIRT to T-Shirt
    if cat_upper in ["POLO T-SHIRT", "POLO T SHIRT", "POLO TSHIRT", "POLO-T-SHIRT"] or cat_upper.startswith("POLO T"):
        return "T-Shirt"
        
    # 2. Convert DENIM to Subcategory
    if cat_upper == "DENIM":
        if subcat_upper in DENIM_SUBCATEGORY_MAP:
            return DENIM_SUBCATEGORY_MAP[subcat_upper]
        if "SHIRT" in subcat_upper:
            return "Shirt"
        if "DUNGAREE" in subcat_upper:
            return "Dungaree"
        if "CO-ORD" in subcat_upper or "COORD" in subcat_upper:
            return "Clothing Set"
        if subcat_upper and subcat_upper not in ["#N/A", "(NIL)", "DENIM", ""]:
            return subcat_upper.title()
        return "Denim"
        
    # 3. Standard category normalizations
    cat_lower = cat_raw.lower()
    if "toy" in cat_lower:
        return "Soft Toys"
    elif "sandal" in cat_lower:
        return "Sandals"
    elif "booties" in cat_lower:
        return "Baby Booties"
        
    return cat_raw.title()

# --- FIELD GENERATORS (BRAND, TAGS, MYNTRA SPECS) ---

def determine_brand(division: str, db: Session) -> str:
    """
    Determines Brand based on Division logic:
    Footwear -> Toothless
    Apparel/Accessories -> Purple United Kids
    """
    div_upper = str(division).upper().strip()
    rule = db.query(BrandRule).filter(BrandRule.division == div_upper).first()
    if rule:
        return rule.brand_name
        
    # Standard Fallback
    if "FOOTWEAR" in div_upper or "SHOE" in div_upper:
        return "Toothless"
    return "Purple United Kids"

def generate_tags(division: str, category: str, gender: str, sub_division: str = "") -> str:
    """
    Generates standard Division, Category, Gender, All Products, NEW LAUNCH tags list.
    Special rules:
    - If gender is Unisex:
      - If sub_division is "Infant", tag is "Infants"
      - If sub_division is anything else, tag is "Unisex"
    - If gender is Infant, tag is "Infants"
    - If gender is Girl / Female, tag is "Girls"
    - If gender is Boy / Male, tag is "Boys"
    - Else fallback to gender title-cased or "Girls"
    """
    div = str(division).title().strip()
    cat = str(category).title().strip()
    gen = str(gender).strip()
    subdiv = str(sub_division).strip()
    
    gen_lower = gen.lower()
    subdiv_lower = subdiv.lower()
    
    # 1. Unisex logic
    if "unisex" in gen_lower:
        if "infant" in subdiv_lower:
            gen_tag = "Infants"
        else:
            gen_tag = "Unisex"
    elif "infant" in gen_lower:
        gen_tag = "Infants"
    elif "girl" in gen_lower or "female" in gen_lower:
        gen_tag = "Girls"
    elif "boy" in gen_lower or "male" in gen_lower:
        gen_tag = "Boys"
    else:
        gen_tag = gen.title() if gen else "Girls"
        
    tags = [div, cat, gen_tag, "All Products", "NEW LAUNCH"]
    return ", ".join(tags)

def generate_myntra_specs(division: str, category: str, gender: str, upper_mat: str, sole_mat: str, fabric_mat: str, item_color: str, db: Session, sub_division: str = "") -> str:
    """
    Generates formatted custom Myntra Specs Info metafield string.
    Uses database-configured templates for category dependency.
    Commodity rules:
    - Unisex Apparel: 'Unisex {Category}' (without 's)
    - Gendered Apparel: 'Girl\\'s {Category}', 'Boy\\'s {Category}'
    """
    div_upper = str(division).upper().strip()
    
    # 1. Fetch template from database
    tmpl = db.query(CategorySpecTemplate).filter(CategorySpecTemplate.division == div_upper).first()
    template_str = tmpl.template_format if tmpl else ""
    
    # Normalizations
    prod_name = str(category).title().strip()
    gender_norm = str(gender).title().strip()
    g_lower = gender_norm.lower()
    
    # Standard formats if template not found or fallback
    if div_upper == "FOOTWEAR":
        # Upper Material, Sole, Items, Commodity
        u = upper_mat or "SYNTHETIC"
        s = sole_mat or "TPR"
        
        if "girl" in g_lower or "female" in g_lower:
            c_gender = "Girls"
        elif "boy" in g_lower or "male" in g_lower:
            c_gender = "Boys"
        elif "unisex" in g_lower:
            c_gender = "Unisex"
        else:
            c_gender = gender_norm or "Unisex"
            
        sub_div_tag = "Infant" if ("booties" in prod_name.lower() or "infant" in str(sub_division).lower() or "infant" in g_lower) else ""
        commodity = f"{c_gender} {sub_div_tag} {prod_name}".replace("  ", " ").strip()
        
        if not template_str:
            template_str = "Item Color: {Item Color}\nUpper Material: {upper}\nSole: {sole}\nItems Included in Packaging: 1 Pair {prod_name}\nCommodity: {commodity}"
            
        return template_str.format(upper=u, sole=s, prod_name=prod_name, commodity=commodity, item_color=item_color, **{"Item Color": item_color})
        
    elif div_upper == "APPAREL":
        # Upper Material (Fabric), Items, Commodity
        f_mat = fabric_mat or upper_mat or "Cotton"
        
        if "girl" in g_lower or "female" in g_lower:
            commodity = f"Girl's {prod_name}"
        elif "boy" in g_lower or "male" in g_lower:
            commodity = f"Boy's {prod_name}"
        elif "unisex" in g_lower:
            commodity = f"Unisex {prod_name}"
        elif "infant" in g_lower:
            commodity = f"Infants {prod_name}" if "infants" in g_lower else f"Infant's {prod_name}"
        else:
            commodity = f"{gender_norm} {prod_name}"
            
        # Decide if "1 Pair" or "1" piece
        pack_term = "1 Pair" if "set" in prod_name.lower() or "suit" in prod_name.lower() or "trouser" in prod_name.lower() else "1"
        
        if not template_str:
            template_str = "Item Color: {Item Color}\nFabric: {fabric}\nItems Included in Packaging: {pack_term} {prod_name}\nCommodity: {commodity}"
            
        return template_str.format(fabric=f_mat, pack_term=pack_term, prod_name=prod_name, commodity=commodity, item_color=item_color, **{"Item Color": item_color})
        
    else:  # ACCESSORIES
        if not template_str:
            template_str = "Item Color: {Item Color}\nItems Included in Packaging: 1 {prod_name}\nCommodity : {prod_name}"
            
        return template_str.format(prod_name=prod_name, item_color=item_color, **{"Item Color": item_color})

# --- CORE VALIDATION LAYER ---

def validate_listings(df_master: pd.DataFrame, master_mappings: dict, df_content: pd.DataFrame, content_mappings: dict) -> list:
    """
    Validates data frames against ecommerce listing requirements.
    Returns a list of warning dicts: { "type": "ERROR"|"WARNING", "message": "...", "sku": "..." }
    Caps returned warnings at 500 to prevent API and frontend freezing.
    """
    warnings = []
    
    # 1. Validate Mastersheet Mappings
    req_master = ["Variant SKU", "Option1 Value", "Variant Compare At Price"]
    for col in req_master:
        if col not in master_mappings:
            warnings.append({"type": "ERROR", "message": f"Mastersheet is missing a column mapped to Shopify '{col}'", "sku": ""})
            
    # 2. Validate Content Mappings
    req_content = ["Title", "Body (HTML)", "Variant Barcode"]
    for col in req_content:
        if col not in content_mappings:
            warnings.append({"type": "ERROR", "message": f"Content Sheet is missing a column mapped to Shopify '{col}'", "sku": ""})
            
    if any(w["type"] == "ERROR" for w in warnings):
        return warnings # Stop validation if core columns missing
        
    # 3. Check for specific records validation
    # Extract headers
    m_sku_col = master_mappings["Variant SKU"]
    m_size_col = master_mappings["Option1 Value"]
    m_price_col = master_mappings["Variant Compare At Price"]
    m_link_col = master_mappings.get("Variant SKU Link", "Item Color")
    m_barcode_col = master_mappings.get("Variant Barcode", "ITEM NAME")
    
    c_style_col = content_mappings["Variant Barcode"]
    c_link_col = content_mappings.get("Variant SKU Link", "SKU")
    c_title_col = content_mappings["Title"]
    
    # Track unique items and build O(1) sets from Content Sheet for lightning-fast matching
    valid_content_links = set(df_content[c_link_col].astype(str).str.strip())
    valid_content_styles = set(df_content[c_style_col].astype(str).str.strip())
    
    seen_skus = set()
    seen_handles = set()
    
    limit_reached = False
    
    # A. Check Mastersheet items
    for idx, row in df_master.iterrows():
        if len(warnings) >= 500:
            limit_reached = True
            break
            
        sku = str(row.get(m_sku_col, "")).strip()
        size = str(row.get(m_size_col, "")).strip()
        price = str(row.get(m_price_col, "")).strip()
        link_val = str(row.get(m_link_col, "")).strip()
        
        # Skip trailing empty rows
        if not sku and not size and not link_val and not price:
            continue
            
        if not sku:
            warnings.append({"type": "ERROR", "message": f"Row {idx+2} in Mastersheet is missing the SKU / ITEM CODE", "sku": ""})
            continue
            
        if sku in seen_skus:
            warnings.append({"type": "WARNING", "message": f"Duplicate Variant SKU/Barcode '{sku}' found in Mastersheet", "sku": sku})
        seen_skus.add(sku)
        
        if not size:
            warnings.append({"type": "WARNING", "message": f"SKU '{sku}' is missing a Size value", "sku": sku})
            
        if not price or price == "0":
            warnings.append({"type": "WARNING", "message": f"SKU '{sku}' has a zero or missing MRP price", "sku": sku})
            
        # Check if Mastersheet item color groups actually match any content rows (O(1) lookups)
        if link_val:
            if link_val not in valid_content_links:
                # Try style code lookup dynamically
                style_val = str(row.get(m_barcode_col, "")).strip()
                if style_val not in valid_content_styles:
                    warnings.append({
                        "type": "WARNING", 
                        "message": f"SKU '{sku}' (color: {link_val}) has no matching record in Content Sheet.", 
                        "sku": sku
                    })

    # Track which style codes and style-color link values are active in df_master
    active_master_links = set(df_master[m_link_col].astype(str).str.strip())
    active_master_styles = set(df_master[m_barcode_col].astype(str).str.strip())

    # B. Check Content Sheet items (only if warnings limit not already reached)
    if not limit_reached:
        for idx, row in df_content.iterrows():
            if len(warnings) >= 500:
                limit_reached = True
                break
                
            style_code = str(row.get(c_style_col, "")).strip()
            title = str(row.get(c_title_col, "")).strip()
            link_val = str(row.get(c_link_col, "")).strip()
            
            # Skip trailing empty rows
            if not style_code and not title and not link_val:
                continue
                
            # Only validate Content Sheet items that are active/present in df_master
            if link_val not in active_master_links and style_code not in active_master_styles:
                continue
                
            if not style_code:
                warnings.append({"type": "WARNING", "message": f"Row {idx+2} in Content Sheet is missing Style/Item Name", "sku": ""})
                
            if not title:
                warnings.append({"type": "ERROR", "message": f"Style '{style_code}' in Content Sheet is missing D2C Title", "sku": style_code})
                
            handle = slugify(title)
            if handle in seen_handles:
                warnings.append({"type": "WARNING", "message": f"Duplicate Handle generated for Title: '{title}'", "sku": style_code})
            seen_handles.add(handle)

            
    if limit_reached:
        warnings.append({
            "type": "WARNING",
            "message": "Validation report truncated: showing first 500 warnings. Please fix these issues first.",
            "sku": ""
        })
        
    return warnings

