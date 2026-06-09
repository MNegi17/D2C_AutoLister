import os
import shutil
import uuid
import time
import pandas as pd
from typing import Dict, List, Optional
from fastapi import FastAPI, UploadFile, File, Form, Depends, HTTPException, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from sqlalchemy.orm import Session

from database import engine, get_db, init_db, ColumnMapping, BrandRule, CategorySpecTemplate, UserCorrection, ProcessingHistory
from core import read_excel_fast, read_csv_fast, auto_map_columns, validate_listings
from generator import generate_shopify_csv

# Setup base workspaces paths dynamically
WORKSPACE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
UPLOAD_DIR = os.path.join(WORKSPACE_DIR, "uploads")
OUTPUT_DIR = os.path.join(WORKSPACE_DIR, "outputs")

os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

def cleanup_old_files(max_age_seconds: int = 7200):
    """
    Clean up temporary upload directories and generated output CSV files older than max_age_seconds (default 2 hours).
    """
    now = time.time()
    
    if os.path.exists(UPLOAD_DIR):
        for item in os.listdir(UPLOAD_DIR):
            item_path = os.path.join(UPLOAD_DIR, item)
            try:
                mtime = os.path.getmtime(item_path)
                if now - mtime > max_age_seconds:
                    if os.path.isdir(item_path):
                        shutil.rmtree(item_path, ignore_errors=True)
                    else:
                        os.remove(item_path)
            except Exception:
                pass
                
    if os.path.exists(OUTPUT_DIR):
        for item in os.listdir(OUTPUT_DIR):
            if item in ["shopify_integration_test_output.csv", "shopify_integration_test_output_populated.csv"]:
                continue
            item_path = os.path.join(OUTPUT_DIR, item)
            try:
                mtime = os.path.getmtime(item_path)
                if now - mtime > max_age_seconds:
                    if os.path.isdir(item_path):
                        shutil.rmtree(item_path, ignore_errors=True)
                    else:
                        os.remove(item_path)
            except Exception:
                pass

app = FastAPI(title="Shopify AutoLister self-learning API")

# Setup CORS for Frontend calls
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Initialize Database on Startup
@app.on_event("startup")
def on_startup():
    init_db()

# --- ENDPOINTS ---

@app.post("/api/upload")
async def upload_files(
    mastersheet: UploadFile = File(...),
    contentsheet: UploadFile = File(...),
    template: UploadFile = File(...),
    historical: Optional[UploadFile] = File(None),
    allowed_item_colors: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    background_tasks: BackgroundTasks = None
):
    if background_tasks:
        background_tasks.add_task(cleanup_old_files)
    """
    Accepts sheets, saves them, runs the auto-mapper, and performs listing validation.
    """
    session_id = str(uuid.uuid4())
    session_upload_dir = os.path.join(UPLOAD_DIR, session_id)
    os.makedirs(session_upload_dir, exist_ok=True)
    
    # Save Uploaded Files
    master_path = os.path.join(session_upload_dir, mastersheet.filename)
    content_path = os.path.join(session_upload_dir, contentsheet.filename)
    template_path = os.path.join(session_upload_dir, template.filename)
    
    with open(master_path, "wb") as buffer:
        shutil.copyfileobj(mastersheet.file, buffer)
    with open(content_path, "wb") as buffer:
        shutil.copyfileobj(contentsheet.file, buffer)
    with open(template_path, "wb") as buffer:
        shutil.copyfileobj(template.file, buffer)
        
    hist_path = None
    if historical:
        hist_path = os.path.join(session_upload_dir, historical.filename)
        with open(hist_path, "wb") as buffer:
            shutil.copyfileobj(historical.file, buffer)
            
    try:
        # Load sheets (Use fast streaming readers)
        df_master = read_excel_fast(master_path, "Report")
        df_content = read_excel_fast(content_path, "MarketplaceD2C")
        df_template = read_csv_fast(template_path)
        
        # Analyze and extract headers
        master_cols = list(df_master.columns)
        content_cols = list(df_content.columns)
        
        # Train system from historical file if provided (Learning Phase)
        if hist_path:
            df_hist = read_csv_fast(hist_path)
            # Find matching column structures between historical CSV and mastersheet/contentsheet
            # Automatically save these column relationships to SQLite ColumnMapping!
            for col in df_hist.columns:
                col_lower = col.lower()
                # Check for direct matches or fuzzy matches to learn relationships
                for m_col in master_cols:
                    if m_col.lower() == col_lower or difflib_ratio(m_col.lower(), col_lower) > 0.85:
                        # Learn Mastersheet relationship
                        existing = db.query(ColumnMapping).filter(ColumnMapping.source_sheet == "mastersheet", ColumnMapping.shopify_column == col).first()
                        if not existing:
                            db.add(ColumnMapping(source_sheet="mastersheet", source_column=m_col, shopify_column=col))
                for c_col in content_cols:
                    if c_col.lower() == col_lower or difflib_ratio(c_col.lower(), col_lower) > 0.85:
                        # Learn Content Sheet relationship
                        existing = db.query(ColumnMapping).filter(ColumnMapping.source_sheet == "contentsheet", ColumnMapping.shopify_column == col).first()
                        if not existing:
                            db.add(ColumnMapping(source_sheet="contentsheet", source_column=c_col, shopify_column=col))
            db.commit()
            
        # Perform Auto Column Mapping
        master_mappings = auto_map_columns(master_cols, "mastersheet", db)
        content_mappings = auto_map_columns(content_cols, "contentsheet", db)
        
        # Apply allowed_item_colors filter to Mastersheet if provided
        if allowed_item_colors and allowed_item_colors.strip():
            import re
            filter_set = {x.strip() for x in re.split(r'[\n\r\t,]+', allowed_item_colors) if x.strip()}
            if filter_set:
                m_link_col = master_mappings.get("Variant SKU Link", "Item Color")
                if m_link_col in df_master.columns:
                    df_master = df_master[df_master[m_link_col].astype(str).str.strip().isin(filter_set)]
        
        # Run validations
        validation_warnings = validate_listings(df_master, master_mappings, df_content, content_mappings)

        
        # Prepare sample data previews
        master_preview = df_master.head(5).to_dict(orient="records")
        content_preview = df_content.head(5).to_dict(orient="records")
        
        return {
            "sessionId": session_id,
            "masterHeaders": master_cols,
            "contentHeaders": content_cols,
            "shopifyHeaders": list(df_template.columns),
            "masterMappings": master_mappings,
            "contentMappings": content_mappings,
            "validationReport": validation_warnings,
            "masterPreview": master_preview,
            "contentPreview": content_preview,
            "files": {
                "mastersheet": master_path,
                "contentsheet": content_path,
                "template": template_path
            }
        }
        
    except Exception as e:
        shutil.rmtree(session_upload_dir, ignore_errors=True)
        raise HTTPException(status_code=500, detail=f"Failed to process uploaded sheets: {str(e)}")

def difflib_ratio(a: str, b: str) -> float:
    import difflib
    return difflib.SequenceMatcher(None, a, b).ratio()

@app.post("/api/mapping/save")
async def save_mappings(
    source_sheet: str = Form(...),
    mappings: str = Form(...),  # JSON string of mappings: { shopify_col: source_col }
    db: Session = Depends(get_db)
):
    """
    Saves user approved column corrections to permanent database memory.
    Overrides future auto mappings.
    """
    import json
    try:
        mapping_dict = json.loads(mappings)
        for shopify_col, source_col in mapping_dict.items():
            if not source_col:
                continue
            # Check if UserCorrection exists, update it, otherwise insert
            existing = db.query(UserCorrection).filter(
                UserCorrection.source_sheet == source_sheet,
                UserCorrection.user_column == shopify_col
            ).first()
            
            if existing:
                existing.source_column = source_col
            else:
                db.add(UserCorrection(
                    source_sheet=source_sheet,
                    source_column=source_col,
                    user_column=shopify_col
                ))
        db.commit()
        return {"status": "success", "message": "User overrides saved permanently."}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Failed to save overrides: {str(e)}")

@app.post("/api/generate")
async def generate_shopify_output(
    session_id: str = Form(...),
    master_path: str = Form(...),
    content_path: str = Form(...),
    template_path: str = Form(...),
    master_mappings: str = Form(...),
    content_mappings: str = Form(...),
    price_overrides: Optional[str] = Form(None), # JSON string of custom price overrides
    allowed_item_colors: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    background_tasks: BackgroundTasks = None
):
    if background_tasks:
        background_tasks.add_task(cleanup_old_files)
    """
    Executes core variant groupings and outputs the Shopify ready CSV listing.
    Records history in SQLite database logs.
    """
    import json
    import re
    try:
        m_mappings = json.loads(master_mappings)
        c_mappings = json.loads(content_mappings)
        p_overrides = json.loads(price_overrides) if price_overrides else {}
        
        # Parse allowed style links if provided
        filter_set = None
        if allowed_item_colors and allowed_item_colors.strip():
            filter_set = {x.strip() for x in re.split(r'[\n\r\t,]+', allowed_item_colors) if x.strip()}
        
        # Load sheets
        df_master = read_excel_fast(master_path, "Report")
        df_content = read_excel_fast(content_path, "MarketplaceD2C")
        df_template = read_csv_fast(template_path)
        
        # Run Generator
        df_output = generate_shopify_csv(
            df_master=df_master,
            master_mappings=m_mappings,
            df_content=df_content,
            content_mappings=c_mappings,
            df_template=df_template,
            db=db,
            price_overrides=p_overrides,
            allowed_style_links=filter_set
        )

        
        # Write Output CSV
        out_filename = f"shopify_upload_{session_id}.csv"
        out_path = os.path.join(OUTPUT_DIR, out_filename)
        
        # Shopify standard requires UTF-8 with BOM for correct Excel rendering of special quotes
        df_output.to_csv(out_path, index=False, encoding="utf-8-sig")
        
        # Calculate stats for historical logs
        total_variants = len(df_output)
        total_products = df_output['Handle'].nunique()
        
        # Save audit history
        history = ProcessingHistory(
            mastersheet_name=os.path.basename(master_path),
            contentsheet_name=os.path.basename(content_path),
            output_csv_name=out_filename,
            total_products=total_products,
            total_variants=total_variants,
            status="Success",
            warnings_count=0
        )
        db.add(history)
        db.commit()
        
        return {
            "status": "success",
            "filename": out_filename,
            "totalProducts": total_products,
            "totalVariants": total_variants,
            "downloadUrl": f"/api/download/{out_filename}"
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to generate listing CSV: {str(e)}")

@app.get("/api/download/{filename}")
async def download_file(filename: str):
    """
    Downloads generated Shopify listings CSVs or populated Matrixify sheets.
    """
    file_path = os.path.join(OUTPUT_DIR, filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Requested file not found.")
    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if filename.endswith(".xlsx") else "text/csv"
    return FileResponse(file_path, media_type=media_type, filename=filename)

@app.get("/api/history")
async def get_history(db: Session = Depends(get_db)):
    """
    Retrieves previous processing histories.
    """
    records = db.query(ProcessingHistory).order_by(ProcessingHistory.timestamp.desc()).all()
    return records

@app.get("/api/learning-status")
async def get_learning_status(db: Session = Depends(get_db)):
    """
    Fetches the size/metrics of saved learnings, rules, overrides in SQLite memory.
    """
    total_mappings = db.query(ColumnMapping).count()
    user_corrections = db.query(UserCorrection).count()
    brand_rules = db.query(BrandRule).count()
    templates = db.query(CategorySpecTemplate).count()
    
    corrections_list = db.query(UserCorrection).all()
    corrections_formatted = [
        {"sheet": c.source_sheet, "original": c.source_column, "target": c.user_column}
        for c in corrections_list
    ]
    
    return {
        "columnMappingsLearned": total_mappings,
        "userCorrectionsSaved": user_corrections,
        "brandRulesConfigured": brand_rules,
        "specTemplatesConfigured": templates,
        "corrections": corrections_formatted
    }

# --- CONFIG RULES CRUD ---

@app.get("/api/brand-rules")
async def get_brand_rules(db: Session = Depends(get_db)):
    return db.query(BrandRule).all()

@app.post("/api/brand-rules/update")
async def update_brand_rule(division: str = Form(...), brand_name: str = Form(...), db: Session = Depends(get_db)):
    rule = db.query(BrandRule).filter(BrandRule.division == division.upper()).first()
    if rule:
        rule.brand_name = brand_name
        db.commit()
        return {"status": "success", "message": f"Brand rule updated for {division}."}
    raise HTTPException(status_code=404, detail=f"Division rule for '{division}' not found.")

@app.get("/api/spec-templates")
async def get_spec_templates(db: Session = Depends(get_db)):
    return db.query(CategorySpecTemplate).all()

@app.post("/api/spec-templates/update")
async def update_spec_template(division: str = Form(...), format_str: str = Form(...), db: Session = Depends(get_db)):
    tmpl = db.query(CategorySpecTemplate).filter(CategorySpecTemplate.division == division.upper()).first()
    if tmpl:
        tmpl.template_format = format_str
        db.commit()
        return {"status": "success", "message": f"Specs template updated for {division}."}
    raise HTTPException(status_code=404, detail=f"Template for '{division}' not found.")

@app.post("/api/matrixify/populate")
async def populate_matrixify_endpoint(
    matrixifyFile: UploadFile = File(...),
    contentsheet: UploadFile = File(...),
    background_tasks: BackgroundTasks = None
):
    if background_tasks:
        background_tasks.add_task(cleanup_old_files)
    """
    Accepts Matrixify and Content Sheet, populates descriptions and bullet points,
    and returns a download link for the updated Matrixify excel sheet.
    """
    session_id = str(uuid.uuid4())
    session_upload_dir = os.path.join(UPLOAD_DIR, f"matrixify_{session_id}")
    os.makedirs(session_upload_dir, exist_ok=True)
    
    mat_path = os.path.join(session_upload_dir, matrixifyFile.filename)
    content_path = os.path.join(session_upload_dir, contentsheet.filename)
    
    with open(mat_path, "wb") as buffer:
        shutil.copyfileobj(matrixifyFile.file, buffer)
    with open(content_path, "wb") as buffer:
        shutil.copyfileobj(contentsheet.file, buffer)
        
    try:
        out_filename = f"populated_{matrixifyFile.filename}"
        out_path = os.path.join(OUTPUT_DIR, out_filename)
        
        # Run population
        import openpyxl
        wb_content = openpyxl.load_workbook(content_path, data_only=True)
        sheet_content = wb_content["MarketplaceD2C"] if "MarketplaceD2C" in wb_content.sheetnames else wb_content.active
        
        rows_content = list(sheet_content.iter_rows(values_only=True))
        headers_content = [str(h).strip() if h is not None else "" for h in rows_content[0]]
        
        sku_idx = -1
        item_name_idx = -1
        title_idx = -1
        desc_idx = -1
        bullets_idx = -1
        
        for idx, h in enumerate(headers_content):
            h_upper = h.upper()
            if h_upper == "SKU":
                sku_idx = idx
            elif h_upper in ["ITEM NAME", "VARIANT BARCODE", "STYLE CODE"]:
                item_name_idx = idx
            elif h_upper in ["D2C TITLE", "TITLE"]:
                title_idx = idx
            elif h_upper in ["DESCRIPTION", "BODY (HTML)"]:
                desc_idx = idx
            elif h_upper in ["BULLET POINTS", "BULLET_POINTS"]:
                bullets_idx = idx
                
        content_by_sku = {}
        content_by_style = {}
        content_by_title = {}
        
        for r in rows_content[1:]:
            if len(r) > max(sku_idx, item_name_idx, title_idx, desc_idx, bullets_idx):
                sku_val = str(r[sku_idx]).strip() if sku_idx != -1 and r[sku_idx] is not None else ""
                style_val = str(r[item_name_idx]).strip() if item_name_idx != -1 and r[item_name_idx] is not None else ""
                title_val = str(r[title_idx]).strip() if title_idx != -1 and r[title_idx] is not None else ""
                desc_val = str(r[desc_idx]).strip() if desc_idx != -1 and r[desc_idx] is not None else ""
                bullets_val = str(r[bullets_idx]).strip() if bullets_idx != -1 and r[bullets_idx] is not None else ""
                
                bullets = []
                for b in bullets_val.split("\n"):
                    b_str = b.strip()
                    if not b_str:
                        continue
                    if b_str.startswith("*"):
                        b_str = "•" + b_str[1:]
                    bullets.append(b_str)
                while len(bullets) < 5:
                    bullets.append("")
                bullets = bullets[:5]
                
                row_data = {
                    "description": desc_val,
                    "bullets": bullets
                }
                
                if sku_val:
                    content_by_sku[sku_val] = row_data
                if style_val:
                    content_by_style[style_val] = row_data
                if title_val:
                    content_by_title[title_val.lower().replace(" ", "")] = row_data
                    
        wb_mat = openpyxl.load_workbook(mat_path)
        sheet_mat = wb_mat.active
        
        headers_mat = [cell.value for cell in sheet_mat[1]]
        
        col_indices = {
            "description": -1,
            "product_info1": -1,
            "product_info2": -1,
            "product_info3": -1,
            "product_info_4": -1,
            "product_info5": -1
        }
        
        for idx, h in enumerate(headers_mat):
            if not h:
                continue
            h_str = str(h)
            if "Variant Metafield: custom.description" in h_str:
                col_indices["description"] = idx
            elif "Variant Metafield: custom.product_info1" in h_str:
                col_indices["product_info1"] = idx
            elif "Variant Metafield: custom.product_info2" in h_str:
                col_indices["product_info2"] = idx
            elif "Variant Metafield: custom.product_info3" in h_str:
                col_indices["product_info3"] = idx
            elif "Variant Metafield: custom.product_info_4" in h_str or "Variant Metafield: custom.product_info4" in h_str:
                col_indices["product_info_4"] = idx
            elif "Variant Metafield: custom.product_info5" in h_str:
                col_indices["product_info5"] = idx
                
        handle_idx = -1
        mat_title_idx = -1
        var_sku_idx = -1
        var_barcode_idx = -1
        
        for idx, h in enumerate(headers_mat):
            if h == "Handle":
                handle_idx = idx
            elif h == "Title":
                mat_title_idx = idx
            elif h == "Variant SKU":
                var_sku_idx = idx
            elif h == "Variant Barcode":
                var_barcode_idx = idx
                
        handle_to_content = {}
        for row_cells in list(sheet_mat.iter_rows(min_row=2)):
            handle_val = row_cells[handle_idx].value if handle_idx != -1 else None
            if not handle_val:
                continue
            handle_str = str(handle_val).strip()
            if handle_str in handle_to_content:
                continue
                
            title_val = row_cells[mat_title_idx].value if mat_title_idx != -1 else None
            sku_val = row_cells[var_sku_idx].value if var_sku_idx != -1 else None
            barcode_val = row_cells[var_barcode_idx].value if var_barcode_idx != -1 else None
            
            matched_data = None
            if barcode_val:
                barcode_str = str(barcode_val).strip()
                matched_data = content_by_style.get(barcode_str)
            if not matched_data and sku_val:
                sku_str = str(sku_val).strip()
                matched_data = content_by_sku.get(sku_str)
                if not matched_data:
                    parts = sku_str.split("-")
                    if len(parts) >= 2:
                        base_sku = "-".join(parts[:2])
                        matched_data = content_by_sku.get(base_sku)
                        if not matched_data:
                            matched_data = content_by_style.get(parts[0])
            if not matched_data and title_val:
                title_clean = str(title_val).lower().replace(" ", "")
                matched_data = content_by_title.get(title_clean)
                if not matched_data:
                    for c_title, data in content_by_title.items():
                        if c_title in title_clean or title_clean in c_title:
                            matched_data = data
                            break
                            
            if matched_data:
                handle_to_content[handle_str] = matched_data
                
        for row_cells in sheet_mat.iter_rows(min_row=2):
            handle_val = row_cells[handle_idx].value if handle_idx != -1 else None
            if not handle_val:
                continue
            handle_str = str(handle_val).strip()
            data = handle_to_content.get(handle_str)
            if not data:
                continue
                
            if col_indices["description"] != -1:
                row_cells[col_indices["description"]].value = data["description"]
            if col_indices["product_info1"] != -1:
                row_cells[col_indices["product_info1"]].value = data["bullets"][0]
            if col_indices["product_info2"] != -1:
                row_cells[col_indices["product_info2"]].value = data["bullets"][1]
            if col_indices["product_info3"] != -1:
                row_cells[col_indices["product_info3"]].value = data["bullets"][2]
            if col_indices["product_info_4"] != -1:
                row_cells[col_indices["product_info_4"]].value = data["bullets"][3]
            if col_indices["product_info5"] != -1:
                row_cells[col_indices["product_info5"]].value = data["bullets"][4]
                
        wb_mat.save(out_path)
        shutil.rmtree(session_upload_dir, ignore_errors=True)
        return {
            "status": "success",
            "filename": out_filename,
            "downloadUrl": f"/api/download/{out_filename}"
        }
    except Exception as e:
        shutil.rmtree(session_upload_dir, ignore_errors=True)
        raise HTTPException(status_code=500, detail=f"Failed to populate Matrixify file: {str(e)}")

if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8080))
    host = os.environ.get("HOST", "0.0.0.0")
    uvicorn.run(app, host=host, port=port)
