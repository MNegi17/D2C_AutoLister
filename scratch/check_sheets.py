import pandas as pd
import openpyxl

mastersheet_path = r"c:\Users\Manann\Desktop\D2C_AutoLister\Sample\ITEM DIRECTORY - MAIN_8June.xlsx"
content_path = r"c:\Users\Manann\Desktop\D2C_AutoLister\Sample\Content Sheet.xlsx"

# Let's inspect the sheets, columns and search for the user's styles.
target_styles = [
    "PUROMS003121-DK. BLUE",
    "PUROMS003120-RED",
    "PBDNSS001233-DK. BLUE",
    "PBDNSS001233-BLUE",
    "PGDRSS002820-LT. BLUE",
    "PBTSHS002792-BROWN",
    "PGDNHS003017-BLUE"
]

print("=== INSPECTING FILES ===")

print("\n1. Reading Content Sheet (first few rows to check structure)...")
wb_c = openpyxl.load_workbook(content_path, read_only=True)
sheet_names_c = wb_c.sheetnames
print("Content Sheet names:", sheet_names_c)
sheet_c = wb_c["MarketplaceD2C"]
c_rows = list(sheet_c.iter_rows(max_row=5, values_only=True))
c_headers = c_rows[0]
print("Content Sheet Headers:", c_headers)

# Find column index for Item Name and SKU
sku_idx = -1
item_name_idx = -1
for idx, h in enumerate(c_headers):
    if str(h).strip().upper() == "SKU":
        sku_idx = idx
    elif str(h).strip().upper() == "ITEM NAME":
        item_name_idx = idx

print(f"SKU index: {sku_idx}, Item Name index: {item_name_idx}")

# Search for matches in Content Sheet
print("\nSearching for target styles in Content Sheet...")
matched_content = []
for row in sheet_c.iter_rows(min_row=2, values_only=True):
    sku_val = str(row[sku_idx]).strip() if sku_idx != -1 and row[sku_idx] is not None else ""
    item_name_val = str(row[item_name_idx]).strip() if item_name_idx != -1 and row[item_name_idx] is not None else ""
    
    # Check if SKU or Item Name contains any of the target styles or part of them
    for ts in target_styles:
        if ts in sku_val or sku_val in ts or ts in item_name_val or item_name_val in ts:
            print(f"  Found row in Content Sheet: SKU='{sku_val}', Item Name='{item_name_val}', D2C Title='{row[2]}'")
            matched_content.append((sku_val, item_name_val))

print("\n2. Reading Mastersheet...")
wb_m = openpyxl.load_workbook(mastersheet_path, read_only=True)
sheet_names_m = wb_m.sheetnames
print("Mastersheet sheet names:", sheet_names_m)
sheet_m = wb_m["Report"]
m_rows = list(sheet_m.iter_rows(max_row=5, values_only=True))
m_headers = m_rows[0]
print("Mastersheet Headers:", m_headers)

# Let's search in Mastersheet
print("\nSearching for target styles in Mastersheet...")
# Find columns like 'Item Color' or 'Variant SKU Link'
item_color_idx = -1
item_name_m_idx = -1
for idx, h in enumerate(m_headers):
    if str(h).strip().upper() == "ITEM COLOR":
        item_color_idx = idx
    elif str(h).strip().upper() == "ITEM NAME":
        item_name_m_idx = idx

print(f"Item Color index: {item_color_idx}, Item Name index: {item_name_m_idx}")

for row in sheet_m.iter_rows(min_row=2, values_only=True):
    color_val = str(row[item_color_idx]).strip() if item_color_idx != -1 and row[item_color_idx] is not None else ""
    item_name_val = str(row[item_name_m_idx]).strip() if item_name_m_idx != -1 and row[item_name_m_idx] is not None else ""
    for ts in target_styles:
        if ts == color_val or ts in color_val or color_val in ts:
            print(f"  Found row in Mastersheet: ITEM NAME='{item_name_val}', ITEM COLOR='{color_val}'")
